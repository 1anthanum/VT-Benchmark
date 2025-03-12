import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelLogger:
    def __init__(self, file_path="results.xlsx", highlight_threshold=90):
        """
        Initialize the Excel Logger.
        
        :param file_path: Path to the Excel file where results will be stored.
        :param highlight_threshold: Test accuracy threshold for highlighting cells.
        """
        self.file_path = file_path
        self.highlight_threshold = highlight_threshold
        self.columns = [
            'Timestamp', 'Dataset', 'Model', 'SMOTE', 'Combination',
            'Optimizer', 'Learning Rate', 'Batch Size',
            'Best Validation Loss', 'Best Validation Accuracy',
            'Final Train Loss', 'Final Train Accuracy',
            'Final Validation Loss', 'Final Validation Accuracy',
            'Test Loss', 'Test Accuracy', 'Precision', 'Recall', 'F1 Score',
            'AUC', 'Balanced Accuracy', 'Specificity'
        ]
        self._initialize_file()

    def _initialize_file(self):
        """Create the Excel file if it does not exist."""
        if not os.path.exists(self.file_path):
            pd.DataFrame(columns=self.columns).to_excel(self.file_path, index=False)

    def log(self, **kwargs):
        """
        Log experiment results to the Excel file.
        
        :param kwargs: Dynamic key-value pairs representing training results.
        Example:
            log(Dataset="ECG5000", Model="Simple2DCNN", Test_Accuracy=92.5)
        """
        kwargs['Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Read existing data or create a new DataFrame
        df = pd.read_excel(self.file_path)
        df = pd.concat([df, pd.DataFrame([kwargs])], ignore_index=True)

        # Save to Excel and apply highlighting
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False)
        self._highlight_high_accuracy()

    def _highlight_high_accuracy(self):
        """Highlight cells where Test Accuracy exceeds the threshold."""
        wb = load_workbook(self.file_path)
        ws = wb.active
        highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Test Accuracy":
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)) and cell.value > self.highlight_threshold:
                        cell.fill = highlight
                break
        wb.save(self.file_path)

    def compute_averages(self):
        """Compute and store the average results of all experiments."""
        df = pd.read_excel(self.file_path)
        avg_results = df.mean(numeric_only=True).to_dict()
        avg_results.update({'Timestamp': 'Average', 'Dataset': 'All', 'Model': 'All'})

        df = pd.concat([df, pd.DataFrame([avg_results])], ignore_index=True)
        df.to_excel(self.file_path, index=False)
        self._highlight_high_accuracy()
        
"""
from excel_output import ExcelLogger

logger = ExcelLogger("results.xlsx")

# Log model results
logger.log(
    Dataset="ECG5000", Model="Simple2DCNN", SMOTE="Yes",
    Combination="Method1", Optimizer="Adam", Learning_Rate=0.001, 
    Batch_Size=64, Best_Validation_Loss=0.12, Best_Validation_Accuracy=91.3,
    Final_Train_Loss=0.09, Final_Train_Accuracy=92.1,
    Final_Validation_Loss=0.11, Final_Validation_Accuracy=90.7,
    Test_Loss=0.10, Test_Accuracy=92.5, Precision=0.91, Recall=0.90,
    F1_Score=0.91, AUC=0.95, Balanced_Accuracy=91.8, Specificity=89.5
)
"""